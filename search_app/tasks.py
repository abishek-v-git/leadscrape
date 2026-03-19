# search_app/tasks.py
import os
import json
import pandas as pd
import requests
import time
from datetime import date
from dotenv import load_dotenv

load_dotenv()

CREDIT_LOG = "media/credit_log.json"


def _load_credit_log():
    if os.path.exists(CREDIT_LOG):
        with open(CREDIT_LOG) as f:
            try:
                return json.load(f)
            except Exception:
                pass
    return []


def _save_credit_entry(entry):
    os.makedirs("media", exist_ok=True)
    log = _load_credit_log()
    log.append(entry)
    # keep last 500 entries
    with open(CREDIT_LOG, "w") as f:
        json.dump(log[-500:], f, indent=2)


def _fetch_pages(base_url, headers, base_payload, target, per_page):
    """Paginate through Apollo results for one search pass. Returns (contacts_list, last_response_headers)."""
    contacts_out = []
    last_headers = {}
    page = 1
    while len(contacts_out) < target:
        payload = {
            **base_payload,
            "page": page,
            "per_page": per_page,
        }
        response = requests.post(
            f"{base_url}/contacts/search",
            json=payload,
            headers=headers,
        )
        last_headers = dict(response.headers)
        if response.status_code != 200:
            print(f"❌ API Error {response.status_code}: {response.text}")
            break
        data = response.json()
        contacts = data.get("contacts", [])
        if not contacts:
            break
        if page == 1 and contacts:
            sample = contacts[0]
            print(f"🔍 RAW FIELD SAMPLE — email={sample.get('email')!r} "
                  f"personal_emails={sample.get('personal_emails')} "
                  f"phone_numbers={sample.get('phone_numbers')} "
                  f"linkedin_url={sample.get('linkedin_url')!r}")
        contacts_out.extend(contacts)
        total_pages = data.get("pagination", {}).get("total_pages", page)
        page += 1
        if page > total_pages:
            break
        time.sleep(0.4)
    return contacts_out, last_headers


def _extract_email(contact):
    # contacts/search returns email directly as a string field.
    return contact.get("email")


def _extract_phone(contact):
    # contacts/search returns phone_numbers as an array of objects.
    numbers = contact.get("phone_numbers") or []
    if numbers:
        return numbers[0].get("sanitized_number") or numbers[0].get("raw_number")
    return contact.get("sanitized_phone")


def _build_entry(contact):
    org = contact.get("organization") or {}
    return {
        "First Name":        contact.get("first_name"),
        "Last Name":         contact.get("last_name"),
        "Job Title":         contact.get("title"),
        "Seniority":         contact.get("seniority"),
        "Email":             _extract_email(contact),
        "Direct Phone":      _extract_phone(contact),
        "LinkedIn Profile":  contact.get("linkedin_url"),
        "Twitter URL":       contact.get("twitter_url"),
        "Facebook URL":      contact.get("facebook_url"),
        "City":              contact.get("city"),
        "State":             contact.get("state"),
        "Country":           contact.get("country"),
        "Headline":          contact.get("headline"),
        "Company Name":      org.get("name"),
        "Industry":          org.get("industry"),
        "Employees":         org.get("estimated_num_employees"),
        "Revenue Range":     org.get("annual_revenue"),
        "Company Website":   org.get("website_url"),
        "Company LinkedIn":  org.get("linkedin_url"),
        "Company Twitter":   org.get("twitter_url"),
        "Company Facebook":  org.get("facebook_url"),
        "Company Phone":     org.get("phone"),
        "Company Description": org.get("description"),
        "Company Addressing": f"{org.get('street_address', '')}, {org.get('city', '')}, {org.get('state', '')} {org.get('postal_code', '')}",
        "SEO Description":   org.get("seo_description"),
        "Founded Year":      org.get("founded_year"),
    }


def search_contacts_sync(search_params):
    print("🚀 STARTING DEEP CONTACT SEARCH")

    api_key = os.getenv("APOLLO_API_KEY")
    if not api_key:
        raise ValueError("APOLLO_API_KEY not found in environment")

    base_url  = "https://api.apollo.io/v1"
    headers   = {"Content-Type": "application/json", "X-Api-Key": api_key}

    # ── Parse inputs ────────────────────────────────────────────────
    keywords     = [k.strip() for k in search_params.get('keywords', '').split(',')    if k.strip()]
    titles       = [t.strip() for t in search_params.get('titles', '').split(',')      if t.strip()]
    seniorities  = [s.strip() for s in search_params.get('seniorities', '').split(',') if s.strip()]
    industries   = [i.strip() for i in search_params.get('industries', '').split(',')  if i.strip()]
    locations    = [l.strip() for l in search_params.get('location', '').split(',')    if l.strip()]
    company_sizes = [cs.strip() for cs in search_params.get('company_size', '').split(',') if cs.strip()]

    max_results = min(int(search_params.get('max_results', 50)), 500)
    per_page    = min(int(search_params.get('per_page', 50)), 100)

    # ── Shared company/seniority filters (always applied) ───────────
    shared = {}
    if seniorities:   shared["person_seniorities"]             = seniorities
    if industries:    shared["organization_industries"]         = industries
    if locations:     shared["person_locations"]               = locations
    if company_sizes: shared["organization_num_employees_ranges"] = company_sizes

    # ── Decide search strategy ───────────────────────────────────────
    raw_contacts = []
    all_resp_headers = {}

    if keywords and titles:
        half = max_results // 2
        print("🔀 Dual-pass: titles pass + keywords pass")
        c1, h1 = _fetch_pages(base_url, headers, {**shared, "person_titles": titles}, half, per_page)
        c2, h2 = _fetch_pages(base_url, headers, {**shared, "q_keywords": " OR ".join(keywords)}, half, per_page)
        raw_contacts = c1 + c2
        all_resp_headers = h2 or h1
    elif titles:
        print("   Single pass — Titles")
        raw_contacts, all_resp_headers = _fetch_pages(base_url, headers, {**shared, "person_titles": titles}, max_results, per_page)
    elif keywords:
        print("   Single pass — Keywords")
        raw_contacts, all_resp_headers = _fetch_pages(base_url, headers, {**shared, "q_keywords": " OR ".join(keywords)}, max_results, per_page)
    else:
        print("   Single pass — filters only")
        raw_contacts, all_resp_headers = _fetch_pages(base_url, headers, shared, max_results, per_page)

    # ── Deduplicate by LinkedIn URL, then email, then full name ─────
    seen  = set()
    results = []
    for contact in raw_contacts:
        uid = (
            contact.get("linkedin_url")
            or contact.get("email")
            or f"{contact.get('first_name')} {contact.get('last_name')} {contact.get('title')}"
        )
        if uid and uid not in seen:
            seen.add(uid)
            results.append(_build_entry(contact))
        if len(results) >= max_results:
            break

    print(f"✅ {len(results)} unique contacts after deduplication")

    # ── Log credit usage ─────────────────────────────────────────────
    revealed = sum(1 for r in results if r.get("Email"))
    _save_credit_entry({
        "date":               str(date.today()),
        "timestamp":          time.strftime("%Y-%m-%d %H:%M:%S"),
        "contacts_fetched":   len(results),
        "emails_revealed":    revealed,
        "rate_limit_limit":   all_resp_headers.get("x-rate-limit-limit", ""),
        "rate_limit_remaining": all_resp_headers.get("x-rate-limit-remaining", ""),
        "hourly_usage":       all_resp_headers.get("x-hourly-usage", ""),
        "daily_usage":        all_resp_headers.get("x-daily-usage", ""),
        "search_titles":      search_params.get("titles", ""),
        "search_keywords":    search_params.get("keywords", ""),
        "search_location":    search_params.get("location", ""),
        "max_results":        max_results,
    })

    # ── Export ───────────────────────────────────────────────────────
    download_dir = "media/downloads"
    if not os.path.exists(download_dir):
        os.makedirs(download_dir)
    else:
        for f in os.listdir(download_dir):
            try: os.remove(os.path.join(download_dir, f))
            except: pass

    df = pd.DataFrame(results)
    df.to_csv(f"{download_dir}/contacts.csv", index=False)

    # Build search filters summary for second sheet
    filters_data = [
        ("Filter",          "Value"),
        ("Keywords",        search_params.get("keywords", "") or "-"),
        ("Job Titles",      search_params.get("titles", "") or "-"),
        ("Seniority",       search_params.get("seniorities", "") or "-"),
        ("Industries",      search_params.get("industries", "") or "-"),
        ("Exclude Industries", search_params.get("exclude_industries", "") or "-"),
        ("Location",        search_params.get("location", "") or "-"),
        ("Company Size",    search_params.get("company_size", "") or "-"),
        ("Max Leads",       str(max_results)),
        ("Total Fetched",   str(len(results))),
        ("Exported On",     time.strftime("%Y-%m-%d %H:%M:%S")),
    ]
    df_filters = pd.DataFrame(filters_data[1:], columns=["Filter", "Value"])

    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        with pd.ExcelWriter(f"{download_dir}/contacts.xlsx", engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Contacts", index=False)
            df_filters.to_excel(writer, sheet_name="Search Filters", index=False)

            ws = writer.sheets["Search Filters"]
            ws.column_dimensions["A"].width = 22
            ws.column_dimensions["B"].width = 55
            header_fill = PatternFill("solid", fgColor="0078D4")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="left")
            for row in ws.iter_rows(min_row=2):
                row[0].font = Font(bold=True)
                row[1].alignment = Alignment(wrap_text=True)
        print("✅ Excel export with Search Filters sheet done")
    except Exception as e:
        print(f"❌ Excel export error: {e}")
        with pd.ExcelWriter(f"{download_dir}/contacts.xlsx", engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Contacts", index=False)
            pd.DataFrame([{"Error": str(e)}]).to_excel(writer, sheet_name="Debug Error", index=False)

    # ── Preview ──────────────────────────────────────────────────────
    preview_data = [
        {
            "Company":            row["Company Name"],
            "Company_Industry":   row["Industry"],
            "Contact_First_Name": row["First Name"],
            "Contact_Last_Name":  row["Last Name"],
            "Contact_Title":      row["Job Title"],
            "Contact_Email":      row["Email"],
            "Contact_LinkedIn":   row["LinkedIn Profile"],
        }
        for row in results
    ]

    return {
        'results_count': len(results),
        'csv_url':       '/media/downloads/contacts.csv',
        'excel_url':     '/media/downloads/contacts.xlsx',
        'preview_data':  preview_data,
    }
